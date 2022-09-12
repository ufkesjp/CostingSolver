# Update_Costs.py corrects costs in spreadsheet with new costs as defined based on matching SKU.

import openpyxl
import pandas as pd

wb = openpyxl.load_workbook('D:\Analytics\Improvement Ideas for Work\Cost_Updates.xlsx')
sheet = wb['Old Costs']
sheet_2 = wb['Cost_Updates']


# Attempt 2:I'm attempting to use the second sheet in the workbook as the dictionary to reference for updates. 
# Trying to convert excel sheet to dictionary
New_Costs = {}
for row in sheet_2.iter_rows(min_row=2, values_only=True):
    SKU2 = row[0]
    Cost = row[1]
    New_Costs[SKU2] = Cost

# To view the dictionary created in the last step
# print(New_Costs)

# Need to loop through the rows, look for matching SKU's, and update the costs.

# Attempt 2: Loop to accomplish the update.
for rowNum in range(2, sheet.max_row): # to skip the first row holding the header
    SKU = sheet.cell(row=rowNum, column = 1).value
    if SKU in New_Costs:
        sheet.cell(row=rowNum, column=2).value = New_Costs[SKU]  
        # I'm currently updating all values to the last value matched rather than changing each one independently.


# Saves updates in a separate workbook in case there are any errors.
wb.save('D:\Analytics\Improvement Ideas for Work\Drafted_Updates.xlsx')