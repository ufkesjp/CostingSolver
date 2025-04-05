# ----------------------------------------
#             Problem Statement
# ----------------------------------------

# In this scenario, a manufacturer needs to determine the cost of individual parts that it purchases in a kit.
#     - The prices of the kit are updated periodically by the supplier. 
#     - The prices of the individual parts are not provided by the supplier, but are necessary for the manufacturer to determine costs of production.
#     - Some parts are used in more than one kit.
#     - Some of the kits have all of the same parts, but are priced differently by the supplier due to slight configuration differences. In instances such as this, the goal is to cost in such a way that minimizes price variance (cost in the system - purchase price).
#
# Question that needs to be answered:
#     What should the cost of each part be updated to after receiving the updated kit prices from the supplier?

# ----------------------------------------
#                Solution
# ----------------------------------------

import openpyxl
import pandas as pd
import numpy as np
import scipy as sp
from scipy.optimize import nnls

# ------------------------------------- Updated ------------------------------------------------

# to read in the composition of each kit, showing which parts are in which kit
KIT_Comp = pd.read_excel('D:\Analytics\Improvement Ideas for Work\Cost_Updates.xlsx', sheet = 'Kit Composition')

# read in the new prices to a separate dataframe - this source serves as the pricing library
KIT_Pricing = pd.read_excel('D:\Analytics\Improvement Ideas for Work\Cost_Updates.xlsx', sheet = 'Kit Prices')

# This is 1-hot encoding the composition, where the quantity indicates whether or not the manufacturer wants the part to receive a portion of the kit price. .getdummies() is another option here.
# The quantity determines the coefficient utilized in the systems of equations non-negative least squares solver. In this case, it's binary (1 or 0).
basket = (KIT_Comp.groupby(['kit_ID', 'part_ID'])['Qty'].sum().unstack().reset_index().fillna(0).set_index('Parent Item)

basket.to_numpy()
A = basket

Price_List = KIT_Pricing ['Cost'].to_numpy()
Price_List = np.nan_to_num(Price_List)

B = Price_List

# utilize scipy's non-negative least squares solver, which prevents negative costs from being assigned to parts.
solution = nnls(A,B)[0]

# Creating a dictionary of new parts costs
COM_List = basket.columns.values

dict = {'part_ID': COM_List, 'Updated Cost': solution}
Updated_COM_Costs = pd.Dataframe(dict)

# previewing this parts dictionary prior to writing it back to excel as a new sheet
print(Updated_Parts_Costs)

# writing the dataframe back into excel - this will throw an error if there's already a sheet named 'Updated_COM_Costs'. This indicates that this work has already been completed.
with pd.ExcelWriter(('D:\Analytics\Improvement Ideas for Work\Cost_Updates.xlsx', sheet = 'Kit Composition', mode = 'a') as writer:
    Updated_COM_Costs.to_excel(writer, sheet_name = 'Updated_COM_Costs')








# ------------------------------------- Previous  ----------------------------------------------
wb = openpyxl.load_workbook('D:\Analytics\Improvement Ideas for Work\Cost_Updates.xlsx')
sheet = wb['Old Costs']
sheet_2 = wb['Cost_Updates']


# Attempt 2:I'm attempting to use the second sheet in the workbook as the dictionary to reference for updates. 
# Trying to convert excel sheet to dictionary
New_Costs = {}
for row in sheet_2.iter_rows(min_row=2, values_only=True):
    SKU2 = row[0]
    Cost = row[1]
    New_Costs[SKU2] = Cost

# To view the dictionary created in the last step
# print(New_Costs)

# Need to loop through the rows, look for matching SKU's, and update the costs.

# Attempt 2: Loop to accomplish the update.
for rowNum in range(2, sheet.max_row): # to skip the first row holding the header
    SKU = sheet.cell(row=rowNum, column = 1).value
    if SKU in New_Costs:
        sheet.cell(row=rowNum, column=2).value = New_Costs[SKU]  
        # I'm currently updating all values to the last value matched rather than changing each one independently.


# Saves updates in a separate workbook in case there are any errors.
wb.save('D:\Analytics\Improvement Ideas for Work\Drafted_Updates.xlsx')
